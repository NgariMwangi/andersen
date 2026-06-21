"""
Kenya payroll run export to Excel (staff payroll items).
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from app.extensions import db
from app.models.consultant import ConsultantPayrollItem
from app.models.employee import Employee
from app.models.payroll import PayrollItem, PayrollRun
from sqlalchemy.orm import joinedload

TWO_DP = Decimal('0.01')
KENYA_NITA_PER_EMPLOYEE = Decimal('50')

# (header label, row dict key) — order defines Excel columns.
KENYA_EXPORT_COLUMNS = [
    ('Employee No.', 'employee_number'),
    ('Employee Name', 'employee_name'),
    ('Job Title', 'job_title'),
    ('Basic Salary', 'basic_salary'),
    ('Benefits', 'benefits'),
    ('Gross Pay', 'gross_pay'),
    ('Taxable Pay', 'taxable_pay'),
    ('SHIF', 'shif'),
    ('Total NSSF', 'total_nssf'),
    ('Welfare Kit', 'welfare_kit'),
    ('SHELLOYEES SACCO', 'shelloyees_sacco'),
    ('MAISHA BORA SACCO', 'maisha_bora_sacco'),
    ('Voluntary Pension', 'voluntary_pension'),
    ('PAYE', 'paye'),
    ('total_deductions', 'total_deductions'),
    ('NET PAY', 'net_pay'),
]

KENYA_EXPORT_HEADERS = [label for label, _ in KENYA_EXPORT_COLUMNS]

TEXT_KEYS = frozenset({'employee_number', 'employee_name', 'job_title'})
NUMERIC_KEYS = tuple(key for _, key in KENYA_EXPORT_COLUMNS if key not in TEXT_KEYS)

# Excel display formats (values stay numeric for sorting/totals).
EXCEL_MONEY_FORMAT = '#,##0.00'
EXCEL_INTEGER_FORMAT = '#,##0'
DEFAULT_MONEY_COLUMN_WIDTH = 16


def autosize_worksheet_columns(
    ws,
    *,
    money_column_indices: set[int] | None = None,
    min_column_widths: dict[int, float] | None = None,
    max_width: float = 44,
) -> None:
    """Size columns from displayed values (comma-formatted amounts need extra width)."""
    from openpyxl.utils import get_column_letter

    money_column_indices = money_column_indices or set()
    min_column_widths = min_column_widths or {}
    for col_idx in range(1, (ws.max_column or 1) + 1):
        max_len = int(min_column_widths.get(col_idx, 8))
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is None or val == '':
                continue
            if col_idx in money_column_indices and isinstance(val, (int, float)):
                display = f'{float(val):,.2f}'
            elif cell.number_format == EXCEL_INTEGER_FORMAT and isinstance(val, (int, float)):
                display = f'{int(val):,}'
            else:
                display = str(val)
            max_len = max(max_len, len(display))
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def min_column_widths_for_headers(
    columns: list[tuple[str, str]],
    text_keys: frozenset[str],
    *,
    text_defaults: dict[int, float] | None = None,
    money_min: float = DEFAULT_MONEY_COLUMN_WIDTH,
) -> dict[int, float]:
    """Minimum widths from header labels and sensible defaults for text columns."""
    text_defaults = text_defaults or {1: 12, 2: 28, 3: 22}
    mins: dict[int, float] = {}
    for idx, (label, key) in enumerate(columns, start=1):
        if key in text_keys:
            mins[idx] = max(text_defaults.get(idx, 10), len(label) + 2)
        else:
            mins[idx] = max(money_min, len(label) + 2)
    return mins

# 1-based column index of each field.
COL_INDEX = {key: idx for idx, (_, key) in enumerate(KENYA_EXPORT_COLUMNS, start=1)}
NUMERIC_COL_INDEX = {key: COL_INDEX[key] for key in NUMERIC_KEYS}
ANALYSIS_LABEL_COL = COL_INDEX['employee_name']

# Deduction line codes excluded when matching recurring/other columns by name.
_STATUTORY_CODES = frozenset({
    'NSSF', 'SHIF', 'HOUSING_LEVY', 'PAYE',
    'PENSION_PERCENT', 'PENSION_FIXED',
})


def _decimal(value) -> Decimal:
    try:
        return Decimal(str(value or 0)).quantize(TWO_DP)
    except Exception:
        return Decimal('0')


def _normalize_name(name: str) -> str:
    return ' '.join((name or '').upper().split())


def _name_matches(name: str, *required_parts: str) -> bool:
    n = _normalize_name(name)
    return all(part.upper() in n for part in required_parts)


def basic_salary_total(item: PayrollItem) -> Decimal:
    """Basic salary from earnings breakdown (BASIC line)."""
    for row in item.earnings_breakdown or []:
        if str(row.get('code') or '').upper() == 'BASIC':
            return _decimal(row.get('amount'))
    return Decimal('0')


def benefits_total(item: PayrollItem) -> Decimal:
    """Sum employee benefit earnings (BEN-* lines) for the period."""
    total = Decimal('0')
    for row in item.earnings_breakdown or []:
        code = str(row.get('code') or '').upper()
        if not code.startswith('BEN-'):
            continue
        total += _decimal(row.get('amount'))
    return total.quantize(TWO_DP)


def _deduction_lines(item: PayrollItem) -> list[dict]:
    return list(item.deductions_breakdown or [])


def _is_statutory_code(code: str) -> bool:
    c = (code or '').upper()
    if c in _STATUTORY_CODES:
        return True
    return c.startswith('NSSF')


def voluntary_pension_total(item: PayrollItem) -> Decimal:
    total = Decimal('0')
    for row in _deduction_lines(item):
        code = (row.get('code') or '').upper()
        if code in ('PENSION_PERCENT', 'PENSION_FIXED'):
            total += _decimal(row.get('amount'))
            continue
        name = row.get('name') or ''
        if _name_matches(name, 'VOLUNTARY', 'PENSION') or _normalize_name(name) == 'VOLUNTARY PENSION':
            total += _decimal(row.get('amount'))
    return total.quantize(TWO_DP)


def _named_deduction_total(item: PayrollItem, *name_parts: str) -> Decimal:
    total = Decimal('0')
    for row in _deduction_lines(item):
        code = (row.get('code') or '').upper()
        if _is_statutory_code(code):
            continue
        if code in ('PENSION_PERCENT', 'PENSION_FIXED'):
            continue
        name = row.get('name') or ''
        if _name_matches(name, *name_parts):
            total += _decimal(row.get('amount'))
    return total.quantize(TWO_DP)


def total_deductions(item: PayrollItem) -> Decimal:
    gross = _decimal(item.gross_pay)
    net = _decimal(item.net_pay)
    return (gross - net).quantize(TWO_DP)


def nssf_employee_employer_total(employee_nssf_total: Decimal) -> Decimal:
    """Employer matches employee NSSF contribution (2× employee total)."""
    return (employee_nssf_total * 2).quantize(TWO_DP)


def kenya_export_row(item: PayrollItem) -> dict:
    emp = item.employee
    job_title = ''
    if emp and emp.job_title:
        job_title = (emp.job_title.name or '').strip()
    return {
        'employee_number': (emp.employee_number or '').strip() if emp else '',
        'employee_name': emp.full_name if emp else f'Employee #{item.employee_id}',
        'job_title': job_title,
        'basic_salary': basic_salary_total(item),
        'benefits': benefits_total(item),
        'gross_pay': _decimal(item.gross_pay),
        'taxable_pay': _decimal(item.taxable_pay),
        'shif': _decimal(item.shif),
        'total_nssf': _decimal(item.nssf_employee),
        'welfare_kit': _named_deduction_total(item, 'WELFARE', 'KIT'),
        'shelloyees_sacco': _named_deduction_total(item, 'SHELLOYEES', 'SACCO'),
        'maisha_bora_sacco': _named_deduction_total(item, 'MAISHA', 'BORA'),
        'voluntary_pension': voluntary_pension_total(item),
        'paye': _decimal(item.paye),
        'total_deductions': total_deductions(item),
        'net_pay': _decimal(item.net_pay),
    }


def _row_to_excel_cells(row: dict) -> list:
    cells = []
    for _, key in KENYA_EXPORT_COLUMNS:
        if key in TEXT_KEYS:
            cells.append(row[key])
        else:
            cells.append(float(row[key]))
    return cells


def _analysis_row(label: str, totals: dict[str, Decimal], *, nssf_emp_employer: Decimal | None = None) -> list:
    """Build an analysis row; optional NSSF (Employee + Employer) only in Total NSSF column."""
    cells = []
    for _, key in KENYA_EXPORT_COLUMNS:
        if key == 'employee_name':
            cells.append(label)
        elif key == 'employee_number':
            cells.append('')
        elif key == 'job_title':
            cells.append('')
        elif key == 'total_nssf' and nssf_emp_employer is not None:
            cells.append(float(nssf_emp_employer))
        elif nssf_emp_employer is not None:
            cells.append('')
        elif key in totals:
            cells.append(float(totals[key]))
        else:
            cells.append('')
    return cells


def _employee_count_row(count: int) -> list:
    cells = []
    for _, key in KENYA_EXPORT_COLUMNS:
        if key == 'employee_name':
            cells.append('Total employees')
        elif key == 'employee_number':
            cells.append('')
        elif key == 'job_title':
            cells.append('')
        elif key == 'basic_salary':
            cells.append(count)
        else:
            cells.append('')
    return cells


def _apply_workbook_number_formats(ws, first_data_row: int) -> None:
    """Comma-separated thousands on all money cells from first data row downward."""
    last_row = ws.max_row
    money_cols = list(NUMERIC_COL_INDEX.values())
    for row_idx in range(first_data_row, last_row + 1):
        label = ws.cell(row=row_idx, column=ANALYSIS_LABEL_COL).value
        for col_idx in money_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None or cell.value == '':
                continue
            if label == 'Total employees' and col_idx == NUMERIC_COL_INDEX['basic_salary']:
                cell.number_format = EXCEL_INTEGER_FORMAT
            elif isinstance(cell.value, (int, float)):
                cell.number_format = EXCEL_MONEY_FORMAT


def _sacco_total(item: PayrollItem) -> Decimal:
    """All SACCO deductions (any line with SACCO in the title)."""
    total = Decimal('0')
    for row in _deduction_lines(item):
        code = (row.get('code') or '').upper()
        if _is_statutory_code(code) or code in ('PENSION_PERCENT', 'PENSION_FIXED'):
            continue
        name = row.get('name') or ''
        if 'SACCO' in _normalize_name(name):
            total += _decimal(row.get('amount'))
    return total.quantize(TWO_DP)


def _helb_total(item: PayrollItem) -> Decimal:
    return _named_deduction_total(item, 'HELB')


def _department_label(item: PayrollItem) -> str:
    emp = item.employee
    if emp and emp.department and (emp.department.name or '').strip():
        return (emp.department.name or '').strip()
    return 'Unassigned'


def gross_pay_by_department(items: list[PayrollItem]) -> list[tuple[str, Decimal]]:
    buckets: dict[str, Decimal] = {}
    for item in items:
        dept = _department_label(item)
        buckets[dept] = buckets.get(dept, Decimal('0')) + _decimal(item.gross_pay)
    return sorted(buckets.items(), key=lambda x: x[0].lower())


def compute_kenya_taxes_summary(
    items: list[PayrollItem],
    consultant_items: list[ConsultantPayrollItem],
) -> dict:
    """Aggregates for the Taxes Summary block (staff + consultants)."""
    dept_gross = gross_pay_by_department(items)
    staff_gross = sum((g for _, g in dept_gross), start=Decimal('0'))
    consultant_gross_rows = []
    consultant_net_rows = []
    wht_total = Decimal('0')
    for idx, ci in enumerate(
        sorted(
            consultant_items,
            key=lambda x: (
                (x.consultant.consultant_number or '') if x.consultant else '',
                x.consultant.full_name if x.consultant else '',
            ),
        ),
        start=1,
    ):
        gross = _decimal(ci.gross_pay)
        net = _decimal(ci.net_pay)
        wht = _decimal(ci.withholding_tax)
        wht_total += wht
        name = ci.consultant.full_name if ci.consultant else f'Consultant #{ci.consultant_id}'
        consultant_gross_rows.append((f'Consultants {idx}', gross, name))
        consultant_net_rows.append((f'Net Pay Consultant {idx}', net, name))

    consultant_gross_total = sum((g for _, g, _ in consultant_gross_rows), start=Decimal('0'))
    staff_net = sum((_decimal(i.net_pay) for i in items), start=Decimal('0'))
    consultant_net_total = sum((n for _, n, _ in consultant_net_rows), start=Decimal('0'))

    return {
        'dept_gross': dept_gross,
        'consultant_gross_rows': consultant_gross_rows,
        'consultant_net_rows': consultant_net_rows,
        'staff_net': staff_net.quantize(TWO_DP),
        'paye': sum((_decimal(i.paye) for i in items), start=Decimal('0')).quantize(TWO_DP),
        'nssf': sum((_decimal(i.nssf_employee) for i in items), start=Decimal('0')).quantize(TWO_DP),
        'nita': (Decimal(len(items)) * KENYA_NITA_PER_EMPLOYEE).quantize(TWO_DP),
        'shif': sum((_decimal(i.shif) for i in items), start=Decimal('0')).quantize(TWO_DP),
        'ahl': sum((_decimal(i.housing_levy) for i in items), start=Decimal('0')).quantize(TWO_DP),
        'wht': wht_total.quantize(TWO_DP),
        'helb': sum((_helb_total(i) for i in items), start=Decimal('0')).quantize(TWO_DP),
        'v_pension': sum((voluntary_pension_total(i) for i in items), start=Decimal('0')).quantize(TWO_DP),
        'sacco': sum((_sacco_total(i) for i in items), start=Decimal('0')).quantize(TWO_DP),
        'welfare': sum((_named_deduction_total(i, 'WELFARE', 'KIT') for i in items), start=Decimal('0')).quantize(TWO_DP),
        'total_gross': (staff_gross + consultant_gross_total).quantize(TWO_DP),
        'total_net': (staff_net + consultant_net_total).quantize(TWO_DP),
        'nssf_employee_employer': nssf_employee_employer_total(
            sum((_decimal(i.nssf_employee) for i in items), start=Decimal('0'))
        ),
    }


def fetch_kenya_consultant_items(run_id: int, company_id: int) -> list[ConsultantPayrollItem]:
    return (
        db.session.query(ConsultantPayrollItem)
        .join(PayrollRun, PayrollRun.id == ConsultantPayrollItem.payroll_run_id)
        .options(joinedload(ConsultantPayrollItem.consultant))
        .filter(
            ConsultantPayrollItem.payroll_run_id == run_id,
            PayrollRun.company_id == company_id,
        )
        .order_by(ConsultantPayrollItem.consultant_id)
        .all()
    )

def fetch_kenya_payroll_items(run_id: int, company_id: int) -> list[PayrollItem]:
    return (
        db.session.query(PayrollItem)
        .join(PayrollRun, PayrollRun.id == PayrollItem.payroll_run_id)
        .options(
            joinedload(PayrollItem.employee).joinedload(Employee.job_title),
            joinedload(PayrollItem.employee).joinedload(Employee.department),
        )
        .filter(
            PayrollItem.payroll_run_id == run_id,
            PayrollRun.company_id == company_id,
        )
        .order_by(PayrollItem.employee_id)
        .all()
    )


def _write_money_cell(ws, row: int, col: int, amount: Decimal | int | float) -> None:
    cell = ws.cell(row=row, column=col, value=float(amount))
    cell.number_format = EXCEL_MONEY_FORMAT


def _append_consultants_section(ws, consultant_items: list[ConsultantPayrollItem]) -> None:
    from openpyxl.styles import Font

    ws.append([])
    header_row = ws.max_row + 1
    ws.append(['CONSULTANTS'])
    ws.cell(row=header_row, column=1).font = Font(bold=True, size=12)

    table_header = ws.max_row + 1
    ws.append([
        'Emp #',
        'Consultants Name',
        'Basic Amount (KES)',
        '',
        '',
        'Gross Pay (KES)',
        '5% WHT',
        'Total to be paid (KES)',
    ])
    for cell in ws[table_header]:
        cell.font = Font(bold=True)

    con_gross = Decimal('0')
    con_wht = Decimal('0')
    con_net = Decimal('0')
    for ci in sorted(
        consultant_items,
        key=lambda x: (
            (x.consultant.consultant_number or '') if x.consultant else '',
            x.consultant.full_name if x.consultant else '',
        ),
    ):
        c = ci.consultant
        gross = _decimal(ci.gross_pay)
        wht = _decimal(ci.withholding_tax)
        net = _decimal(ci.net_pay)
        con_gross += gross
        con_wht += wht
        con_net += net
        ws.append([
            (c.consultant_number or '').strip() if c else '',
            c.full_name if c else f'Consultant #{ci.consultant_id}',
            float(gross),
            '',
            '',
            float(gross),
            float(wht),
            float(net),
        ])

    if consultant_items:
        total_row = ws.max_row + 1
        ws.append([
            'Grand Total',
            len(consultant_items),
            float(con_gross),
            '',
            '',
            float(con_gross),
            float(con_wht),
            float(con_net),
        ])
        for cell in ws[total_row]:
            cell.font = Font(bold=True)
        for col in (3, 6, 7, 8):
            ws.cell(row=total_row, column=col).number_format = EXCEL_MONEY_FORMAT
        ws.cell(row=total_row, column=2).number_format = EXCEL_INTEGER_FORMAT

    for row_idx in range(table_header + 1, ws.max_row + 1):
        for col in (3, 6, 7, 8):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = EXCEL_MONEY_FORMAT


def _append_taxes_summary(ws, summary: dict, employee_count: int) -> None:
    from openpyxl.styles import Font, PatternFill

    ws.append([])
    header_row = ws.max_row + 1
    ws.cell(row=header_row, column=1, value='Taxes Summary')
    title_cell = ws.cell(row=header_row, column=1)
    title_cell.font = Font(bold=True, color='FFFFFF', size=12)
    title_cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')

    def summary_line(label: str, amount: Decimal | None = None, *, bold: bool = False) -> None:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=label)
        if amount is not None:
            _write_money_cell(ws, r, 2, amount)
        if bold:
            ws.cell(row=r, column=1).font = Font(bold=True)
            if amount is not None:
                ws.cell(row=r, column=2).font = Font(bold=True)

    for dept_name, gross in summary['dept_gross']:
        summary_line(dept_name, gross)

    for label, gross, _name in summary['consultant_gross_rows']:
        summary_line(f'{label}', gross)

    ws.append([])
    summary_line('Net Pay Employees', summary['staff_net'])
    for label, net, _name in summary['consultant_net_rows']:
        summary_line(label, net)

    ws.append([])
    summary_line('PAYE', summary['paye'])
    summary_line('NSSF', summary['nssf'])
    summary_line('NITA', summary['nita'])
    summary_line('SHIF', summary['shif'])
    summary_line('AHL', summary['ahl'])
    summary_line('WHT', summary['wht'])
    summary_line('HELB', summary['helb'])
    summary_line('V.Pension', summary['v_pension'])
    summary_line('SACCO', summary['sacco'])
    summary_line('WELFARE', summary['welfare'])

    ws.append([])
    summary_line('Total Gross Pay', summary['total_gross'], bold=True)
    summary_line('Total Net Pay', summary['total_net'], bold=True)
    summary_line('NSSF (Employee + Employer)', summary['nssf_employee_employer'], bold=True)
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value='Total employees')
    ws.cell(row=r, column=2, value=employee_count)
    ws.cell(row=r, column=2).number_format = EXCEL_INTEGER_FORMAT


def build_kenya_payroll_workbook(
    run: PayrollRun,
    items: list[PayrollItem],
    consultant_items: list[ConsultantPayrollItem] | None = None,
) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    consultant_items = consultant_items or []

    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {run.pay_year}-{run.pay_month:02d}"

    ws.append(KENYA_EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    totals = {k: Decimal('0') for k in NUMERIC_KEYS}
    first_data_row = 2

    for item in items:
        row = kenya_export_row(item)
        ws.append(_row_to_excel_cells(row))
        for k in NUMERIC_KEYS:
            totals[k] += row[k]

    employee_count = len(items)

    # Per-column totals directly under the employee table
    ws.append([])
    analysis_header_row = ws.max_row + 1
    ws.append(['ANALYSIS'] + [''] * (len(KENYA_EXPORT_HEADERS) - 1))
    for cell in ws[analysis_header_row]:
        cell.font = Font(bold=True, size=12)

    ws.append(_employee_count_row(employee_count))
    ws.append(_analysis_row('Total (all employees)', totals))
    nssf_combined = nssf_employee_employer_total(totals['total_nssf'])
    ws.append(
        _analysis_row(
            'NSSF (Employee + Employer)',
            totals,
            nssf_emp_employer=nssf_combined,
        )
    )

    for row_idx in range(analysis_header_row + 1, ws.max_row + 1):
        for cell in ws[row_idx]:
            cell.font = Font(bold=True)

    if consultant_items:
        _append_consultants_section(ws, consultant_items)

    summary = compute_kenya_taxes_summary(items, consultant_items)
    _append_taxes_summary(ws, summary, employee_count)

    _apply_workbook_number_formats(ws, first_data_row)

    ws.freeze_panes = 'D2'
    autosize_worksheet_columns(
        ws,
        money_column_indices=set(NUMERIC_COL_INDEX.values()),
        min_column_widths=min_column_widths_for_headers(KENYA_EXPORT_COLUMNS, TEXT_KEYS),
    )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
