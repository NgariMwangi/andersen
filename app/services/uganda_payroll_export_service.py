"""
Uganda payroll run export to Excel (staff payroll items).
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from app.extensions import db
from app.models.employee import Employee
from app.models.payroll import PayrollItem, PayrollRun
from sqlalchemy.orm import joinedload

from app.services.kenya_payroll_export_service import (
    EXCEL_INTEGER_FORMAT,
    EXCEL_MONEY_FORMAT,
    autosize_worksheet_columns,
    benefits_total,
    basic_salary_total,
    min_column_widths_for_headers,
    total_deductions,
    _decimal,
    _named_deduction_total,
)

TWO_DP = Decimal('0.01')

UGANDA_EXPORT_COLUMNS = [
    ('Employee No.', 'employee_number'),
    ('Employee Name', 'employee_name'),
    ('Job Title', 'job_title'),
    ('Basic Salary', 'basic_salary'),
    ('Benefits', 'benefits'),
    ('Gross Pay', 'gross_pay'),
    ('PAYE', 'paye'),
    ('NSSF', 'nssf'),
    ('Welfare Kit', 'welfare_kit'),
    ('Total Deductions', 'total_deductions'),
    ('Net Pay', 'net_pay'),
    ('NSSF Employer', 'nssf_employer'),
    ('Total Payroll Cost', 'total_payroll_cost'),
]

UGANDA_EXPORT_HEADERS = [label for label, _ in UGANDA_EXPORT_COLUMNS]

TEXT_KEYS = frozenset({'employee_number', 'employee_name', 'job_title'})
NUMERIC_KEYS = tuple(key for _, key in UGANDA_EXPORT_COLUMNS if key not in TEXT_KEYS)

COL_INDEX = {key: idx for idx, (_, key) in enumerate(UGANDA_EXPORT_COLUMNS, start=1)}
NUMERIC_COL_INDEX = {key: COL_INDEX[key] for key in NUMERIC_KEYS}
ANALYSIS_LABEL_COL = COL_INDEX['employee_name']
FREEZE_PANES = 'D2'


def total_payroll_cost(item: PayrollItem) -> Decimal:
    """Employer cost: gross pay plus employer NSSF contribution."""
    return (_decimal(item.gross_pay) + _decimal(item.nssf_employer)).quantize(TWO_DP)


def uganda_export_row(item: PayrollItem) -> dict:
    emp = item.employee
    job_title = ''
    if emp and emp.job_title:
        job_title = (emp.job_title.name or '').strip()
    return {
        'employee_number': (emp.employee_number or '').strip() if emp else '',
        'employee_name': emp.full_name if emp else f'Employee #{item.employee_id}',
        'job_title': job_title,
        'basic_salary': basic_salary_total(item),
        'benefits': benefits_total(item),
        'gross_pay': _decimal(item.gross_pay),
        'paye': _decimal(item.paye),
        'nssf': _decimal(item.nssf_employee),
        'welfare_kit': _named_deduction_total(item, 'WELFARE', 'KIT'),
        'total_deductions': total_deductions(item),
        'net_pay': _decimal(item.net_pay),
        'nssf_employer': _decimal(item.nssf_employer),
        'total_payroll_cost': total_payroll_cost(item),
    }


def _row_to_excel_cells(row: dict) -> list:
    cells = []
    for _, key in UGANDA_EXPORT_COLUMNS:
        if key in TEXT_KEYS:
            cells.append(row[key])
        else:
            cells.append(float(row[key]))
    return cells


def _analysis_row(label: str, totals: dict[str, Decimal]) -> list:
    cells = []
    for _, key in UGANDA_EXPORT_COLUMNS:
        if key == 'employee_name':
            cells.append(label)
        elif key in TEXT_KEYS:
            cells.append('')
        elif key in totals:
            cells.append(float(totals[key]))
        else:
            cells.append('')
    return cells


def _employee_count_row(count: int) -> list:
    cells = []
    for _, key in UGANDA_EXPORT_COLUMNS:
        if key == 'employee_name':
            cells.append('Total employees')
        elif key == 'basic_salary':
            cells.append(count)
        else:
            cells.append('')
    return cells


def _apply_workbook_number_formats(ws, first_data_row: int) -> None:
    from openpyxl.styles import Alignment

    last_row = ws.max_row
    money_cols = list(NUMERIC_COL_INDEX.values())
    money_align = Alignment(horizontal='right')
    for row_idx in range(first_data_row, last_row + 1):
        label = ws.cell(row=row_idx, column=ANALYSIS_LABEL_COL).value
        for col_idx in money_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None or cell.value == '':
                continue
            if label == 'Total employees' and col_idx == NUMERIC_COL_INDEX['basic_salary']:
                cell.number_format = EXCEL_INTEGER_FORMAT
            elif isinstance(cell.value, (int, float)):
                cell.number_format = EXCEL_MONEY_FORMAT
                cell.alignment = money_align


def fetch_uganda_payroll_items(run_id: int, company_id: int) -> list[PayrollItem]:
    return (
        db.session.query(PayrollItem)
        .join(PayrollRun, PayrollRun.id == PayrollItem.payroll_run_id)
        .options(joinedload(PayrollItem.employee).joinedload(Employee.job_title))
        .filter(
            PayrollItem.payroll_run_id == run_id,
            PayrollRun.company_id == company_id,
        )
        .order_by(PayrollItem.employee_id)
        .all()
    )


def build_uganda_payroll_workbook(run: PayrollRun, items: list[PayrollItem]) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {run.pay_year}-{run.pay_month:02d}"

    ws.append(UGANDA_EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[1].height = 32

    totals = {k: Decimal('0') for k in NUMERIC_KEYS}
    first_data_row = 2

    for item in items:
        row = uganda_export_row(item)
        ws.append(_row_to_excel_cells(row))
        for k in NUMERIC_KEYS:
            totals[k] += row[k]

    employee_count = len(items)

    ws.append([])
    analysis_header_row = ws.max_row + 1
    ws.append(['ANALYSIS'] + [''] * (len(UGANDA_EXPORT_HEADERS) - 1))
    for cell in ws[analysis_header_row]:
        cell.font = Font(bold=True, size=12)

    ws.append(_employee_count_row(employee_count))
    ws.append(_analysis_row('Total (all employees)', totals))

    for row_idx in range(analysis_header_row + 1, ws.max_row + 1):
        for cell in ws[row_idx]:
            cell.font = Font(bold=True)

    _apply_workbook_number_formats(ws, first_data_row)
    ws.freeze_panes = FREEZE_PANES

    autosize_worksheet_columns(
        ws,
        money_column_indices=set(NUMERIC_COL_INDEX.values()),
        min_column_widths=min_column_widths_for_headers(UGANDA_EXPORT_COLUMNS, TEXT_KEYS),
    )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
